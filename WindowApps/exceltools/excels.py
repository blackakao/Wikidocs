import openpyxl

# 김포점 오전보고서 긁어와보기, 엑셀 시트 만들어보기, 해당열로 접근 가능한지 조종해보기 등등

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection

# 새로운 엑셀 인스턴스 생성
# wb = openpyxl.Workbook()
# 생성 후 워크시트 1개 자동 생성 된다고 함
# ws  = wb.active
# 활성화 된 워크시트를 가리킴

# 기존의 엑셀파일 로드
wb = openpyxl.load_workbook(filename='C:\\Users\\User\\Desktop\\JB\\본사\\시설현황\\2. 지점별문서\\2022\\202201\\20220125\\(김포점)입소자&근로자 현황보고(20220125).xlsx')
# 엑셀파일 로드
ws = wb.active
# 활성시트 가리키기
ws = wb['22.02']
# 특정 시트 가리키기

# wb.create_sheet('New_sheet',0)
# 시트 명 = '새시트이름', 0번째(0=맨 왼쪽부터) 시트를 위치시키겠음

# print(ws['A1']) #A1 셀 자체
# print(ws['A1'].value) #A1 셀의 값
# 화면에서 디버깅 하고 싶을 때 

# print(ws.cell(row = 5, column= 4).value)
# 5행 4열에 있는 값을 가져오겠다

# print(ws['B'].value)

for cell in ws['D']:
    print(cell.value)
# D열에 있는 모든 값을 가지고 오겠다

# for cell in ws['D']:
#     cell.font = Font(color="0099CC00")
# D열에 있는 글자색을 구리구리하게 바꾸어 보겠다


wb.save(filename='C:\\Users\\User\\Desktop\\JB\\본사\\시설현황\\2. 지점별문서\\2022\\202201\\20220125\\(김포점)입소자&근로자 현황보고(20220125).xlsx')
# 가지고 논 엑셀 파일 저장

wb.close()
# 엑셀 종료



